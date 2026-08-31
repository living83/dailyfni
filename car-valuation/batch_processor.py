"""Batch valuate an Excel of loan records.

Reads via `excel_loader.load_loans`, fans out to `valuation_engine.valuate`
under an asyncio.Semaphore, writes an enriched copy with the original sheet
preserved + 6 extra columns appended.
"""
from __future__ import annotations

import asyncio
import logging
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, SCRAPER_CONCURRENCY
from excel_loader import LoanRow, load_loans
from valuation_engine import SCRAPER_CLASSES, valuate

logger = logging.getLogger(__name__)

# Columns we append to the workbook (in this order).
OUTPUT_COLUMNS: tuple[str, ...] = (
    "추정시세",
    "공매처분가",
    "엔카매물수",
    "KB매물수",
    "헤이딜러매물수",
    "신뢰도등급",
)


@dataclass(slots=True)
class BatchSummary:
    input_path: Path
    output_path: Path
    total_rows: int
    processed: int
    succeeded: int      # confidence != "실패"
    failed: int         # confidence == "실패"
    elapsed_sec: float
    warnings: list[str]


def _default_output_path(input_path: Path) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = input_path.stem
    return OUTPUT_DIR / f"{stem}__valued.xlsx"


async def _valuate_with_sem(
    sem: asyncio.Semaphore,
    loan: LoanRow,
    *,
    use_cache: bool,
) -> tuple[LoanRow, dict[str, Any]]:
    async with sem:
        result = await valuate(loan.model, loan.year, use_cache=use_cache)
    return loan, result


def _row_to_output_values(result: dict[str, Any]) -> dict[str, Any]:
    """Map a valuate() result onto the 6 output column values."""
    sources = result.get("sources") or {}
    def site_count(name: str) -> int | str:
        d = sources.get(name)
        return d["count"] if d else 0
    return {
        "추정시세": result.get("market_price"),
        "공매처분가": result.get("auction_price"),
        "엔카매물수": site_count("encar"),
        "KB매물수": site_count("kbchachacha"),
        "헤이딜러매물수": site_count("heydealer"),
        "신뢰도등급": result.get("confidence"),
    }


def _write_output_xlsx(
    input_path: Path,
    output_path: Path,
    header_row_offset: int,
    results_by_row: dict[int, dict[str, Any]],
) -> None:
    """Copy the input sheet to output_path, append 6 new columns."""
    wb = load_workbook(filename=input_path)
    ws = wb.active
    if ws is None:
        wb.close()
        raise RuntimeError(f"no active sheet in {input_path}")

    header_excel_row = header_row_offset + 1  # 1-based
    last_col = ws.max_column
    # Write new headers
    for j, name in enumerate(OUTPUT_COLUMNS, start=last_col + 1):
        ws.cell(row=header_excel_row, column=j, value=name)

    # Write per-row values
    for excel_row_num, values in results_by_row.items():
        for j, name in enumerate(OUTPUT_COLUMNS, start=last_col + 1):
            ws.cell(row=excel_row_num, column=j, value=values.get(name))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def _detect_header_offset(input_path: Path) -> int:
    """Best-effort: rerun the same header detection used by excel_loader so
    we know where to place the new column headers."""
    from excel_loader import detect_header_row
    wb = load_workbook(filename=input_path, data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return detect_header_row(rows)


async def valuate_batch(
    input_path: Path | str,
    output_path: Path | str | None = None,
    *,
    use_cache: bool = True,
    concurrency: int | None = None,
    column_overrides: dict[str, str] | None = None,
) -> BatchSummary:
    """End-to-end Excel → valuate → enriched Excel."""
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else _default_output_path(input_path)
    started = time.monotonic()

    loans, load_warnings = load_loans(input_path, column_overrides=column_overrides)
    logger.info("loaded %d loans from %s (%d skipped)", len(loans), input_path, len(load_warnings))

    sem = asyncio.Semaphore(concurrency or SCRAPER_CONCURRENCY)
    tasks = [_valuate_with_sem(sem, loan, use_cache=use_cache) for loan in loans]

    results_by_row: dict[int, dict[str, Any]] = {}
    succeeded = 0
    failed = 0
    for coro in asyncio.as_completed(tasks):
        loan, result = await coro
        results_by_row[loan.row_idx] = _row_to_output_values(result)
        if result.get("confidence") == "실패":
            failed += 1
        else:
            succeeded += 1
        logger.info(
            "row %d %s %d → %s (canonical=%s)",
            loan.row_idx, loan.model, loan.year,
            result.get("confidence"), result.get("canonical"),
        )

    header_offset = _detect_header_offset(input_path)
    _write_output_xlsx(input_path, output_path, header_offset, results_by_row)

    return BatchSummary(
        input_path=input_path,
        output_path=output_path,
        total_rows=len(loans) + len(load_warnings),
        processed=len(loans),
        succeeded=succeeded,
        failed=failed,
        elapsed_sec=round(time.monotonic() - started, 2),
        warnings=load_warnings,
    )
