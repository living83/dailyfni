"""Read an NPL loan ledger Excel file and emit clean per-row records.

NPL operators send ledgers with inconsistent column naming ("차종" vs "차량명",
"연식" vs "차량연식" vs "년식"). We fuzzy-match headers against a small known
set and let the user override via the `column_map` argument.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

logger = logging.getLogger(__name__)


# Header aliases. The key is the canonical column we emit; the value lists
# alternative spellings the operator might use.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "model": ("차종", "차량명", "차종명", "모델명", "모델", "차량종류"),
    "year": ("연식", "년식", "차량연식", "출고연식", "제조연도", "year"),
    "plate": ("차량번호", "번호판", "차번호", "등록번호"),
}

# Header detection scoring threshold (0-100). Below this we leave the column
# unmapped and surface it as a warning.
HEADER_MATCH_THRESHOLD = 80


@dataclass(slots=True)
class LoanRow:
    row_idx: int                 # 1-based Excel row number (so row 2 is the first data row when header is row 1)
    model: str
    year: int
    plate: str | None
    raw: dict[str, Any]          # all original columns, header-keyed


def _norm_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    # Drop punctuation/whitespace; case-fold latin parts
    s = re.sub(r"[\s\-_()\[\]:.]+", "", s)
    return s.lower()


def detect_header_row(rows: list[list[Any]], *, max_scan: int = 5) -> int:
    """Find the row index (0-based) that looks most like a header.

    NPL ledgers sometimes have a title row above the actual header. We score
    each of the first `max_scan` rows by how many cells contain header-like
    Korean text and return the highest-scoring one.
    """
    flat_aliases = {a for aliases in HEADER_ALIASES.values() for a in aliases}
    norm_aliases = {_norm_header(a) for a in flat_aliases}

    best_i = 0
    best_score = -1
    for i, row in enumerate(rows[:max_scan]):
        hits = sum(
            1 for c in row
            if c is not None and _norm_header(str(c)) in norm_aliases
        )
        if hits > best_score:
            best_score = hits
            best_i = i
    return best_i


def map_columns(
    headers: list[str | None],
    *,
    overrides: dict[str, str] | None = None,
    threshold: int = HEADER_MATCH_THRESHOLD,
) -> dict[str, int]:
    """Return canonical_key → column_index (0-based) mapping.

    `overrides` lets the operator pin a column by header text, e.g.
    {"model": "차종_지칭"}.
    """
    normalized_headers = [_norm_header(h) for h in headers]
    mapping: dict[str, int] = {}

    for canonical, aliases in HEADER_ALIASES.items():
        # Explicit override?
        if overrides and canonical in overrides:
            target = _norm_header(overrides[canonical])
            for j, h in enumerate(normalized_headers):
                if h == target:
                    mapping[canonical] = j
                    break
            if canonical in mapping:
                continue

        # Exact alias hit first
        norm_aliases = [_norm_header(a) for a in aliases]
        found = False
        for j, h in enumerate(normalized_headers):
            if h in norm_aliases:
                mapping[canonical] = j
                found = True
                break
        if found:
            continue

        # Fuzzy fallback per alias
        best: tuple[int, float] | None = None
        for j, h in enumerate(normalized_headers):
            if not h:
                continue
            score = max((fuzz.ratio(h, a) for a in norm_aliases), default=0)
            if best is None or score > best[1]:
                best = (j, score)
        if best and best[1] >= threshold:
            mapping[canonical] = best[0]

    return mapping


def _parse_year(value: Any) -> int | None:
    """Coerce a cell value into a 4-digit year. Accepts 2014, "2014", "14",
    "14년", "2014년식", datetime objects, etc.
    """
    if value is None:
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2100 else (2000 + value if 0 <= value < 70 else None)
    # datetime → year
    if hasattr(value, "year"):
        return int(value.year)
    s = str(value).strip()
    m = re.search(r"(\d{4})", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d{2})", s)
    if m:
        yy = int(m.group(1))
        return 2000 + yy if yy < 70 else 1900 + yy
    return None


def load_loans(
    path: Path | str,
    *,
    sheet: str | int | None = None,
    column_overrides: dict[str, str] | None = None,
) -> tuple[list[LoanRow], list[str]]:
    """Read the workbook and return (rows, warnings).

    `rows` only contains data rows where both `model` and `year` resolved.
    Skipped rows go into the warnings list with their row number + reason.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else (
        wb.worksheets[sheet] if isinstance(sheet, int) else wb.active
    )

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        wb.close()
        return [], [f"sheet {ws.title!r} is empty"]

    header_idx = detect_header_row([list(r) for r in all_rows])
    header_row = list(all_rows[header_idx])
    headers = [None if h is None else str(h).strip() for h in header_row]

    mapping = map_columns(headers, overrides=column_overrides)
    if "model" not in mapping or "year" not in mapping:
        wb.close()
        missing = [k for k in ("model", "year") if k not in mapping]
        raise ValueError(
            f"required column(s) not found: {missing}. "
            f"Detected headers: {headers}. "
            f"Use column_overrides to pin them, e.g. {{'model': '차종', 'year': '연식'}}"
        )

    warnings_: list[str] = []
    loans: list[LoanRow] = []
    for excel_row_num, raw in enumerate(all_rows[header_idx + 1 :], start=header_idx + 2):
        # excel_row_num is 1-based and aligns with what the user sees in Excel
        raw_list = list(raw)
        if all(c is None or str(c).strip() == "" for c in raw_list):
            continue  # blank row

        raw_dict: dict[str, Any] = {}
        for j, h in enumerate(headers):
            key = h if h else f"col{get_column_letter(j+1)}"
            raw_dict[key] = raw_list[j] if j < len(raw_list) else None

        model_cell = raw_list[mapping["model"]] if mapping["model"] < len(raw_list) else None
        year_cell = raw_list[mapping["year"]] if mapping["year"] < len(raw_list) else None
        plate_cell = (
            raw_list[mapping["plate"]]
            if "plate" in mapping and mapping["plate"] < len(raw_list)
            else None
        )

        model = (str(model_cell).strip() if model_cell is not None else "")
        year = _parse_year(year_cell)
        if not model:
            warnings_.append(f"row {excel_row_num}: empty 차종 — skipped")
            continue
        if year is None:
            warnings_.append(f"row {excel_row_num}: unparseable 연식 ({year_cell!r}) — skipped")
            continue

        loans.append(
            LoanRow(
                row_idx=excel_row_num,
                model=model,
                year=year,
                plate=(str(plate_cell).strip() if plate_cell else None),
                raw=raw_dict,
            )
        )

    wb.close()
    return loans, warnings_
