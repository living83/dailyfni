from __future__ import annotations

import asyncio
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import batch_processor as bp_module  # noqa: E402
from batch_processor import OUTPUT_COLUMNS, valuate_batch  # noqa: E402


def _make_input_xlsx(path: Path, rows: list[list]) -> Path:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


@pytest.fixture
def stub_valuate(monkeypatch):
    """Per-row stub: matches K7 (canonical), fails for others."""
    async def _stub(model: str, year: int, *, use_cache: bool = True, **_):
        if "K7" in model.upper() or "케이7" in model:
            return {
                "model": model, "canonical": "K7", "year": year,
                "market_price": 6_316_000, "auction_price": 4_737_000,
                "confidence": "상",
                "sources": {
                    "encar":       {"median": 5_990_000, "count": 20},
                    "kbchachacha": {"median": 6_700_000, "count": 9},
                    "heydealer":   {"median": 6_300_000, "count": 3},
                },
                "cached": True,
                "cache_hits": {"encar": True, "kbchachacha": True, "heydealer": True},
                "elapsed_sec": 0.01, "match_score": 100.0,
            }
        return {
            "model": model, "canonical": None, "year": year,
            "market_price": None, "auction_price": None,
            "confidence": "실패",
            "sources": {"encar": None, "kbchachacha": None, "heydealer": None},
            "cached": False,
            "cache_hits": {"encar": False, "kbchachacha": False, "heydealer": False},
            "elapsed_sec": 0.01, "match_score": 0.0,
        }
    monkeypatch.setattr(bp_module, "valuate", _stub)
    return _stub


@pytest.mark.asyncio
async def test_batch_writes_appended_columns_and_summary(stub_valuate, tmp_path: Path):
    in_path = _make_input_xlsx(tmp_path / "in.xlsx", [
        ["채권번호", "담보종류", "차종", "연식", "차량번호", "미상환원금잔액"],
        ["A001", "자동차", "K7", 2014, "12가3456", 8_000_000],
        ["A002", "자동차", "기아 K7", 2015, "34나5678", 7_000_000],
        ["A003", "자동차", "완전미상차종XYZ", 2010, "78다9012", 5_000_000],
    ])
    out_path = tmp_path / "out.xlsx"

    summary = await valuate_batch(in_path, out_path, use_cache=False, concurrency=2)

    assert summary.processed == 3
    assert summary.succeeded == 2     # K7, 기아 K7
    assert summary.failed == 1        # 완전미상차종XYZ
    assert out_path.exists()

    wb = load_workbook(out_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for col in OUTPUT_COLUMNS:
        assert col in headers, f"output column {col!r} missing; got {headers}"

    # Find column indices for new columns
    col_idx = {h: i for i, h in enumerate(headers, start=1) if h in OUTPUT_COLUMNS}

    # Row 2 (K7) → "상", 추정시세 6,316,000
    row2 = {h: ws.cell(row=2, column=i).value for h, i in col_idx.items()}
    assert row2["신뢰도등급"] == "상"
    assert row2["추정시세"] == 6_316_000
    assert row2["공매처분가"] == 4_737_000
    assert row2["엔카매물수"] == 20
    assert row2["KB매물수"] == 9
    assert row2["헤이딜러매물수"] == 3

    # Row 4 (unknown) → "실패", prices None
    row4 = {h: ws.cell(row=4, column=i).value for h, i in col_idx.items()}
    assert row4["신뢰도등급"] == "실패"
    assert row4["추정시세"] is None
    assert row4["공매처분가"] is None
    assert row4["엔카매물수"] == 0
    wb.close()


@pytest.mark.asyncio
async def test_concurrency_cap_enforced(stub_valuate, tmp_path: Path, monkeypatch):
    """Patch valuate to track max simultaneous calls; must not exceed limit."""
    in_path = _make_input_xlsx(tmp_path / "in.xlsx", [
        ["차종", "연식"],
        *[["K7", 2014] for _ in range(12)],
    ])

    counter = {"current": 0, "peak": 0}
    base_stub = bp_module.valuate
    async def _tracked(model, year, **kw):
        counter["current"] += 1
        counter["peak"] = max(counter["peak"], counter["current"])
        await asyncio.sleep(0.05)
        try:
            return await base_stub(model, year, **kw)
        finally:
            counter["current"] -= 1
    monkeypatch.setattr(bp_module, "valuate", _tracked)

    await valuate_batch(in_path, tmp_path / "out.xlsx", concurrency=3)
    assert counter["peak"] <= 3, f"peak concurrency {counter['peak']} exceeded cap 3"
    assert counter["peak"] >= 2, "concurrency 3 should have parallelized at least 2"
